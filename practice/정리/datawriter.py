import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

# wb = openpyxl.load_workbook('test.xlsx')
# WS = wb.active
#
# WS_1 = wb.create_sheet('두번째 시트')

# WS["A1"] = "difficult"

df1 = pd.read_csv('iphone.csv', index_col=0)
df2 = pd.read_csv('iphone.csv', index_col=0)


with pd.ExcelWriter('test.xlsx', mode='a', engine='openpyxl') as writer:
    df1.to_excel(writer, sheet_name="c")


# writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
# df1.to_excel(writer, sheet_name = "a" )
# df2.to_excel(writer, sheet_name = "b" )

# wb.save('test.xlsx')