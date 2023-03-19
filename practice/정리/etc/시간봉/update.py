import pandas as pd
import ccxt
import openpyxl as xl
from datetime import datetime

'''api 키 '''
with open("../api(JKBOT).txt") as f:
    lines = f.readlines()
    api_key = lines[0].strip()
    secret = lines[1].strip()


'''바이낸스 접속'''
binance = ccxt.binance(config={
    'apiKey': api_key, #"XTpKrQGSk3GhXzqiEV4OfwGJzmTVcLh8dKGwHo4aQBH4p0mOqPDpIsxdh95tjGVf",
    'secret': secret, #"A0eqZGEWHsyL3NMM6WuDrucIanr7A2YZAnrwXVPhXpf2WGauIANwa5zsoNeNt0hs",
    'enableRateLimit' : True,
    'options': {
        'defaultType': 'future'
    }
})
jongmok=[]
markets = binance.load_markets()
for market in markets.keys():
    if market.endswith("/USDT"):
        jongmok.append(market)

# wb = xl.load_workbook('database.xlsx')
a = 0

for symbol in jongmok:
    a = a + 1
    print(a, symbol)
    try:
        wb = xl.load_workbook('Database/' + symbol[:-5] + '.xlsx')
    except:
        btc_ohlcv = binance.fetch_ohlcv(symbol=symbol, timeframe='1d', since=None, limit=100)
        df = pd.DataFrame(data=btc_ohlcv, columns=["datetime", "open", "high", "low", "close", "volumns"])
        df['datetime'] = pd.to_datetime(df['datetime'], unit='ms')
        # df = df.astype({'datetime': 'str'})
        df.set_index('datetime', inplace=True)
        df.set_index('datetime', inplace=True)

        df.to_excel('Database/' + symbol[:-5] + '.xlsx', sheet_name=symbol[:-5])


    btc_ohlcv = binance.fetch_ohlcv(symbol=symbol, timeframe='1d', since=None, limit=100)
    df = pd.DataFrame(data=btc_ohlcv, columns=["datetime", "open", "high", "low", "close", "volumns"])
    df['datetime'] = pd.to_datetime(df['datetime'], unit='ms')
    # df = df.astype({'datetime': 'str'})
    df.set_index('datetime', inplace=True)

    df.to_excel('test.xlsx', sheet_name='가격정보')

    new_wb = xl.load_workbook('test.xlsx')
    new_ws = new_wb['가격정보']
    df = pd.DataFrame(new_ws.values, columns=["datetime", "open", "high", "low", "close", "volumns"])
    df.drop(0, axis='index', inplace=True)
    # df.set_index('datetime', inplace=True)

    ws = wb[symbol[:-5]]
    before_data = pd.DataFrame(ws.values, columns=["datetime", "open", "high", "low", "close", "volumns"])

    before_data.drop(0, axis='index', inplace=True)
    # before_data = before_data.astype({'datetime': 'str'})
    # before_data['datetime'] = before_data['datetime'].str[:-9]
    #

    print(before_data)
    print(df)

    after_data = pd.concat([before_data, df], axis=0, join='outer')
    after_data = after_data.drop_duplicates(['datetime'], keep='first')
    after_data.set_index('datetime', inplace=True)

    print(after_data)
    after_data.to_excel('Database/' + symbol[:-5] + '.xlsx', sheet_name=symbol[:-5])

    # with pd.ExcelWriter('database.xlsx', engine='openpyxl', if_sheet_exists='replace') as writer:
    #     df.to_excel(writer, index=False, sheet_name=symbol[:-5])

#

# for name in jongmok:#[:data_count]:
#     btc_ohlcv = binance.fetch_ohlcv(name, timeframe='1d', since=None, limit=data_day)
#     print(name)
#
#     #print(btc_ohlcv)
#     df = pd.DataFrame(btc_ohlcv, columns = ['datetime', 'open', 'high', 'low', 'close', 'volumn'])
#     df['datetime'] = pd.to_datetime(df['datetime'], unit='ms')
#     df.set_index('datetime', inplace=True)
#     print(df)
#
#     with pd.ExcelWriter(file_name, mode='a', engine='openpyxl') as writer:
#         df.to_excel(writer, sheet_name=name[:-5])