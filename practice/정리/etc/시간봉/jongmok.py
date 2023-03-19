import ccxt
import pprint

from openpyxl import Workbook
from openpyxl import load_workbook

#file_name 생성될 파일
#base_file 기초 매크로 파일
def Create_file(base_file, file_name, data_count):

    binance = ccxt.binance()
    markets = binance.load_markets()
    jongmok = []
    new_jongmok = []

    for market in markets.keys():
        if market.endswith("/USDT"):
            jongmok.append(market)

    for name in jongmok[:data_count]:
        new_jongmok.append(name[:-5])
    print(new_jongmok)
    #
    WB = load_workbook(base_file)
    # WS = WB["통계"]
    # # WS_1 = WB.create_sheet()

    a = 1
    # for i in new_jongmok:
    #     WS.cell(row=9, column=a+2, value=i)
    #     a = a + 1
    #
    WB.save(file_name)