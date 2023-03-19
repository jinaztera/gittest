import openpyxl as xl
import pandas as pd

from datetime import datetime


# wb1 = xl.load_workbook("Test1.xlsx")
# wb2 = xl.load_workbook("Test2.xlsx")
#
# print(len(wb1.sheetnames))
#

# df1 = pd.read_excel('Test1.xlsx')
# print(df1)
# df2 = pd.read_excel('Test2.xlsx')
# print(df2)
#
# df3 = df1.merge(df2, how="outer")
#
# wb = xl.load_workbook('Test1.xlsx')
#
# df1 = df3
# print(df3.iloc[:, 0:5])
#
# print(add(3, 2))
#
# def add(a, b):
#     c = a + b
#     return c
#

# time = 2021-1-1
# type(2021-1-1)
# print(time)
# for i in range(10):
a = 0
def sublist_max(profits):
    # 코드를 작성하세요.
    # maxsum = profits[0]  # 부분최대값'
    # largesum = profits[0]  # 최대값
    for j in range(len(profits)):
        maxsum = profits[j]

        for i in range(len(profits) - j - 1):
            if maxsum < a + profits[i + j]:
                maxsum = a + profits[i + j + 1]

            a = a + profits[i + 1]

        if maxsum >= largesum:
            largesum = maxsum

    return largesum


# 테스트
print(sublist_max([4, 3, 8, -2, -5, -3, -5, -3]))
print(sublist_max([2, 3, 1, -1, -2, 5, -1, -1]))
print(sublist_max([7, -3, 14, -8, -5, 6, 8, -5, -4, 10, -1, 8]))